#!/usr/bin/env python3
# =============================================================================
# generate_data.py — Lyzr Pipeline data normaliser
# =============================================================================
# Reads the consolidated tracker spreadsheet (xlsx) and writes pipeline/data.json
# with the SAME schema + aggregates the live Cloudflare Pages function
# (functions/api/rows.js -> recomputeAggregates) produces.
#
# Dependency: openpyxl  (read-only xlsx parsing, no native build step).
#   pip install openpyxl
#
# Usage:
#   python3 generate_data.py <input.xlsx> <output.json>
#   python3 generate_data.py --help
#
# On success it prints a one-line summary:
#   row count, total ACV, unique companies, owner-leaderboard size.
#
# Normalisation rules (mirrors pipeline/README.md):
#   - "a" in any field            -> blank (None / "")
#   - owner cells split on , and / -> discrete people
#   - canonicalise Bharath        -> Bharath Bhat
#   - split Joel/Mark             -> Joel Kandy, Mark Leibowitz
#   - ACV "$500,000"              -> 500000 (acv), original string kept as acv_raw
#   - Deal Close Date             -> close_date_raw preserved; close_quarter derived
#   - Prototype Link              -> only kept/clickable if a valid http(s) URL
# =============================================================================

import argparse
import json
import re
import sys
from datetime import datetime, timezone

# ---------------------------------------------------------------------------
# Configuration: segment prefixes, allowed stages, owner canonicalisation.
# ---------------------------------------------------------------------------

SEGMENT_PREFIX = {
    "Internal": "INT",
    "Accenture": "ACC",
    "GSI-SI": "GSI",
    "Enterprises": "ENT",
}

# Normalise a handful of segment spellings seen in source sheets.
SEGMENT_ALIASES = {
    "internal": "Internal",
    "accenture": "Accenture",
    "gsi-si": "GSI-SI",
    "gsi/si": "GSI-SI",
    "gsi": "GSI-SI",
    "si": "GSI-SI",
    "enterprise": "Enterprises",
    "enterprises": "Enterprises",
    "enterprise sales": "Enterprises",
}

# Canonical stage labels used throughout the dashboard.
STAGE_ALIASES = {
    "demo": "Demo",
    "in-conversation": "in-conversation",
    "in conversation": "in-conversation",
    "inconversation": "in-conversation",
    "conversation": "in-conversation",
    "win": "win",
    "won": "win",
    "customer": "customer",
    "lost": "lost",
    "loss": "lost",
}

# Owner-name canonicalisation. A single source token on the LEFT maps to one or
# more discrete owner names on the RIGHT. Tokens that split into multiple people
# (e.g. "Joel/Mark") are expanded here AFTER the / and , splitting, so they only
# fire when the token is exactly that combined name.
OWNER_CANONICAL = {
    "bharath": ["Bharath Bhat"],
    "joel/mark": ["Joel Kandy", "Mark Leibowitz"],
    "joel": ["Joel Kandy"],
    "mark": ["Mark Leibowitz"],
}

URL_RE = re.compile(r"^https?://", re.IGNORECASE)

# The "a" placeholder (case-insensitive, trimmed) means "blank" in the source.
BLANK_TOKENS = {"a", "n/a", "na", "-", "--", "tbd", "."}


# ---------------------------------------------------------------------------
# Cell-level normalisation
# ---------------------------------------------------------------------------

def clean_text(value):
    """Return a trimmed string, or None for blanks / the "a" placeholder."""
    if value is None:
        return None
    s = str(value).strip()
    if s == "" or s.lower() in BLANK_TOKENS:
        return None
    return s


def parse_owners(value):
    """Split an owner cell on ',' and '/' into discrete, canonicalised people.

    Order is preserved and duplicates are removed (first occurrence wins).
    """
    text = clean_text(value)
    if not text:
        return []

    # Split on commas and slashes. We first try to honour the explicit
    # "Joel/Mark" combined token before generic slash-splitting, so it maps to
    # the two named people rather than to "Joel" + "Mark" generically. (Both
    # routes reach the same result here, but keeping the combined entry makes
    # the intent from the README explicit and future-proofs other combos.)
    raw_parts = re.split(r"[,/]", text)
    out = []
    for part in raw_parts:
        token = part.strip()
        if not token:
            continue
        key = token.lower()
        if key in OWNER_CANONICAL:
            names = OWNER_CANONICAL[key]
        else:
            names = [token]
        for name in names:
            if name not in out:
                out.append(name)
    return out


def parse_acv(value):
    """Return (acv_number_or_None, acv_raw_string_or_None).

    acv keeps the numeric value for sorting/aggregation; acv_raw preserves the
    original cell text (e.g. "$500,000") for display.
    """
    if value is None:
        return None, None

    # Numeric cells come straight through.
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        num = float(value)
        if num != num:  # NaN guard
            return None, None
        num = int(num) if num.is_integer() else num
        return num, str(value)

    raw = str(value).strip()
    cleaned = clean_text(raw)
    if not cleaned:
        return None, None

    # Strip currency symbols, thousands separators and whitespace, keep digits
    # and a single decimal point.
    digits = re.sub(r"[^0-9.]", "", cleaned)
    if digits in ("", "."):
        return None, raw  # non-numeric text: preserve raw, no number
    try:
        num = float(digits)
    except ValueError:
        return None, raw
    num = int(num) if num.is_integer() else num
    return num, raw


def derive_close_quarter(close_date_raw):
    """Derive a 'Q<n> <year>' label from a raw close-date string where possible.

    Handles:
      - already-quarter strings: "Q2 2026", "q3 2025"
      - common date strings:     "2/25/2026", "2026-02-25", "25/02/2026",
                                 "Feb 2026", "February 25, 2026"
    Returns None when nothing parseable is found.
    """
    text = clean_text(close_date_raw)
    if not text:
        return None

    # Already a quarter label, e.g. "Q2 2026" or "Q2-2026".
    m = re.search(r"\bq\s*([1-4])\D*((?:19|20)\d{2})\b", text, re.IGNORECASE)
    if m:
        return f"Q{m.group(1)} {m.group(2)}"

    # Numeric date: M/D/YYYY or M-D-YYYY (US-style, month first).
    m = re.search(r"\b(\d{1,2})[/\-](\d{1,2})[/\-]((?:19|20)?\d{2})\b", text)
    if m:
        month = int(m.group(1))
        year = int(m.group(3))
        if year < 100:
            year += 2000
        if 1 <= month <= 12:
            return f"Q{(month - 1) // 3 + 1} {year}"

    # ISO date: YYYY-MM-DD.
    m = re.search(r"\b((?:19|20)\d{2})[/\-](\d{1,2})[/\-](\d{1,2})\b", text)
    if m:
        year = int(m.group(1))
        month = int(m.group(2))
        if 1 <= month <= 12:
            return f"Q{(month - 1) // 3 + 1} {year}"

    # Month name + year, e.g. "Feb 2026" or "February 25, 2026".
    months = {
        "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
        "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
    }
    m = re.search(r"\b([A-Za-z]{3,9})\b.*?\b((?:19|20)\d{2})\b", text)
    if m:
        key = m.group(1)[:3].lower()
        if key in months:
            return f"Q{(months[key] - 1) // 3 + 1} {int(m.group(2))}"

    return None


def parse_prototype_link(value):
    """Return (link_or_None, link_text_or_None).

    Only a valid http(s) URL is kept and marked clickable; anything else is
    dropped so the dashboard never renders a broken link.
    """
    text = clean_text(value)
    if not text:
        return None, None
    if URL_RE.match(text):
        return text, text
    return None, None


def normalise_segment(value):
    text = clean_text(value)
    if not text:
        return None
    return SEGMENT_ALIASES.get(text.lower(), text)


def normalise_stage(value):
    text = clean_text(value)
    if not text:
        return None
    return STAGE_ALIASES.get(text.lower(), text)


def normalise_industry(value):
    return clean_text(value)


# ---------------------------------------------------------------------------
# Column mapping: tolerate header variations from the source sheet.
# ---------------------------------------------------------------------------

# Each logical field maps to a list of acceptable header spellings (lower-cased,
# non-alphanumerics stripped). First match wins.
COLUMN_SYNONYMS = {
    "sn": ["sn", "sno", "serialno", "serial", "no", "slno"],
    "segment": ["segment", "category2", "vertical", "track", "pod"],
    "company": ["company", "companyname", "account", "accountname", "client", "customer"],
    "industry": ["industry", "sector", "vertical2"],
    "project": ["project", "projectname", "opportunity", "opportunityname", "dealname", "deal"],
    "use_case": ["usecase", "usecases", "usecase2"],
    "category": ["category", "type", "solution", "solutiontype"],
    "stage": ["stage", "status", "dealstage"],
    "prototype_owners": ["prototypeowner", "prototypeowners", "protoowner", "protoowners", "builtby", "demoowner"],
    "opportunity_owners": ["opportunityowner", "opportunityowners", "oppowner", "oppowners", "owner", "owners", "dealowner", "salesowner"],
    "prototype_link": ["prototypelink", "prototype", "demolink", "link", "prototypeurl", "demourl"],
    "acv": ["acv", "annualcontractvalue", "contractvalue", "value", "dealvalue", "amount"],
    "time_period": ["timeperiod", "period", "contractterm", "term", "duration"],
    "close_date_raw": ["closedate", "dealclosedate", "expectedclosedate", "closedaterow", "close"],
}


def _norm_header(h):
    return re.sub(r"[^a-z0-9]", "", str(h or "").lower())


def build_column_index(header_row):
    """Map logical field -> column index based on header synonyms."""
    norm_headers = [_norm_header(h) for h in header_row]
    index = {}
    for field, synonyms in COLUMN_SYNONYMS.items():
        for col_idx, nh in enumerate(norm_headers):
            if nh and nh in synonyms:
                index[field] = col_idx
                break
    return index


# ---------------------------------------------------------------------------
# Aggregates + facets — a faithful Python port of recomputeAggregates() in
# functions/api/rows.js so the output is byte-compatible with the live API.
# ---------------------------------------------------------------------------

def recompute_aggregates(rows):
    total_acv = 0
    acv_open = acv_won = acv_lost = 0
    by_stage = {}
    by_segment = {}
    company_map = {}
    owner_map = {}
    unique_companies = set()

    for r in rows:
        acv = r.get("acv") or 0
        total_acv += acv
        if r.get("stage") == "lost":
            acv_lost += acv
        elif r.get("stage") in ("win", "customer"):
            acv_won += acv
        else:
            acv_open += acv

        stage = r.get("stage")
        by_stage[stage] = by_stage.get(stage, 0) + 1

        seg = r.get("segment")
        if seg not in by_segment:
            by_segment[seg] = {"count": 0, "acv": 0, "stages": {}}
        by_segment[seg]["count"] += 1
        by_segment[seg]["acv"] += acv
        by_segment[seg]["stages"][stage] = by_segment[seg]["stages"].get(stage, 0) + 1

        company = r.get("company")
        if company:
            unique_companies.add(company)
            if company not in company_map:
                company_map[company] = {"company": company, "deals": 0, "acv": 0}
            company_map[company]["deals"] += 1
            company_map[company]["acv"] += acv

        for o in r.get("opportunity_owners") or []:
            if o not in owner_map:
                owner_map[o] = {"name": o, "deals": 0, "acv": 0}
            owner_map[o]["deals"] += 1
            owner_map[o]["acv"] += acv

    # Stable sort matching the JS comparator: acv desc, then deals desc.
    owner_leaderboard = sorted(
        owner_map.values(), key=lambda x: (-x["acv"], -x["deals"])
    )
    top_companies = sorted(
        company_map.values(), key=lambda x: (-x["acv"], -x["deals"])
    )

    aggregates = {
        "total_rows": len(rows),
        "total_acv": total_acv,
        "acv_open": acv_open,
        "acv_won": acv_won,
        "acv_lost": acv_lost,
        "unique_companies": len(unique_companies),
        "by_stage": by_stage,
        "by_segment": by_segment,
        "owner_leaderboard": owner_leaderboard,
        "top_companies": top_companies,
    }

    facets = {
        "segments": sorted({r["segment"] for r in rows if r.get("segment")}),
        "stages": list(dict.fromkeys(r["stage"] for r in rows if r.get("stage"))),
        "categories": sorted({r["category"] for r in rows if r.get("category")}),
        "industries": sorted({
            "BFSI" if (r.get("industry") or "").startswith("BFSI") else r["industry"]
            for r in rows if r.get("industry")
        }),
    }

    return aggregates, facets


# ---------------------------------------------------------------------------
# Row building
# ---------------------------------------------------------------------------

def build_row(raw, segment, sn):
    """Build one normalised row dict in the canonical key order."""
    acv, acv_raw = parse_acv(raw.get("acv"))
    link, link_text = parse_prototype_link(raw.get("prototype_link"))
    close_date_raw = clean_text(raw.get("close_date_raw"))

    project = clean_text(raw.get("project"))
    use_case = clean_text(raw.get("use_case")) or project

    return {
        "id": None,  # filled in by caller (needs full-rows context)
        "segment": segment,
        "sn": sn,
        "company": clean_text(raw.get("company")),
        "industry": normalise_industry(raw.get("industry")),
        "project": project,
        "use_case": use_case,
        "category": clean_text(raw.get("category")),
        "stage": normalise_stage(raw.get("stage")),
        "prototype_owners": parse_owners(raw.get("prototype_owners")),
        "opportunity_owners": parse_owners(raw.get("opportunity_owners")),
        "prototype_link": link,
        "prototype_link_text": link_text,
        "acv": acv,
        "acv_raw": acv_raw,
        "time_period": clean_text(raw.get("time_period")),
        "close_date_raw": close_date_raw,
        "close_quarter": derive_close_quarter(close_date_raw),
        "edit_history": [],
    }


def next_id(segment, max_num):
    prefix = SEGMENT_PREFIX.get(segment, "ROW")
    return f"{prefix}-{str(max_num).zfill(4)}"


# ---------------------------------------------------------------------------
# Spreadsheet reading
# ---------------------------------------------------------------------------

def read_rows(xlsx_path):
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit(
            "ERROR: openpyxl is required. Install it with:\n"
            "  pip install openpyxl"
        )

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    raw_rows = []

    for ws in wb.worksheets:
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            continue  # empty sheet

        col = build_column_index(header)
        # A usable sheet must at least identify a project/opportunity column.
        if "project" not in col and "company" not in col:
            continue

        # If the sheet has no Segment column, fall back to the sheet name as the
        # segment (a common per-tab layout: one tab per segment).
        sheet_segment = normalise_segment(ws.title)

        for values in rows_iter:
            if values is None:
                continue
            # Skip fully blank rows.
            if all(v is None or str(v).strip() == "" for v in values):
                continue

            def cell(field):
                idx = col.get(field)
                if idx is None or idx >= len(values):
                    return None
                return values[idx]

            raw = {field: cell(field) for field in COLUMN_SYNONYMS}
            seg = normalise_segment(raw.get("segment")) or sheet_segment
            if not seg:
                continue  # cannot place this row in a segment
            # Require at least a project or a company to be a real row.
            if not clean_text(raw.get("project")) and not clean_text(raw.get("company")):
                continue
            raw_rows.append((raw, seg))

    wb.close()
    return raw_rows


def generate(xlsx_path, source_file_label):
    raw_rows = read_rows(xlsx_path)

    rows = []
    sn = 1
    max_num = 0
    for raw, seg in raw_rows:
        row = build_row(raw, seg, sn)
        # default stage if the sheet omitted it, so aggregates stay clean
        if not row["stage"]:
            row["stage"] = "in-conversation"
        max_num += 1
        row["id"] = next_id(seg, max_num)
        rows.append(row)
        sn += 1

    aggregates, facets = recompute_aggregates(rows)

    return {
        "generated_at": datetime.now(timezone.utc)
        .isoformat(timespec="milliseconds")
        .replace("+00:00", "Z"),
        "source_file": source_file_label,
        "rows": rows,
        "aggregates": aggregates,
        "facets": facets,
    }


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main(argv=None):
    parser = argparse.ArgumentParser(
        description="Normalise the pipeline tracker xlsx into data.json."
    )
    parser.add_argument("input", help="Path to the source tracker .xlsx")
    parser.add_argument("output", help="Path to write data.json")
    parser.add_argument(
        "--source-label",
        default=None,
        help="Value stored in data.json 'source_file' (default: input basename)",
    )
    args = parser.parse_args(argv)

    import os
    label = args.source_label or os.path.basename(args.input)

    data = generate(args.input, label)

    with open(args.output, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
        f.write("\n")

    agg = data["aggregates"]
    print(
        f"Rows: {agg['total_rows']}  |  "
        f"Total ACV: ${agg['total_acv']:,}  |  "
        f"Unique companies: {agg['unique_companies']}  |  "
        f"Owners on leaderboard: {len(agg['owner_leaderboard'])}"
    )


if __name__ == "__main__":
    main()
